# -*- coding: utf-8 -*-
"""通用化 xlsx 生成脚本(schema 驱动,领域无关).

来源:从 output/build_xlsx.py 通用化而来,移除 7 张血液专表硬编码。
"""
from __future__ import annotations
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = Font(bold=True, color="FFFFFF", name="Microsoft YaHei", size=11)
NAVY_FILL = PatternFill("solid", fgColor="003366")
BLUE_FILL = PatternFill("solid", fgColor="E8F0FE")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

XLSX_MIN_SIZE_BYTES = 1024  # 1 KB · build_xlsx size_ok 阈值
XLSX_MIN_SHEET_COUNT = 1


def _style_header(ws: Any) -> None:
    for cell in ws[1]:
        cell.font = NAVY
        cell.fill = NAVY_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _style_body(ws: Any) -> None:
    max_r, max_c = ws.max_row, ws.max_column
    for r in range(2, max_r + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(r, c)
            cell.alignment = LEFT if c >= 3 else CENTER
            cell.border = BORDER
            cell.font = Font(name="Microsoft YaHei", size=10)
            if r % 2 == 0:
                cell.fill = BLUE_FILL


def _autofit(ws: Any, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _validate_xlsx(out_path: Path, sheet_count: int) -> dict:
    size = out_path.stat().st_size
    return {
        "size_ok": size > XLSX_MIN_SIZE_BYTES,
        "size_bytes": size,
        "sheet_count": sheet_count,
        "sheet_count_ok": sheet_count >= XLSX_MIN_SHEET_COUNT,
    }


def build_xlsx(
    data: dict[str, list[list[Any]]],
    schema_yaml: Path,
    out_path: Path,
    *,
    auto_validate: bool = True,
) -> dict:
    """根据 schema YAML 生成多 sheet xlsx。

    Args:
        data: {data_key: [[row1...], [row2...]]} · key 与 schema.sheets[].data_key 对应
        schema_yaml: schema 文件,格式见 docs(sheets / headers / column_widths / data_key)
        out_path: 输出 xlsx 路径
        auto_validate: 自动验证(size > XLSX_MIN_SIZE_BYTES,sheet_count >= XLSX_MIN_SHEET_COUNT)

    Returns:
        {"ok", "out_path", "size_mb", "sheets": [...], "validation"}

    Raises:
        FileNotFoundError: schema_yaml 不存在
        ValueError: schema 缺 'sheets' 字段或为空
        KeyError: schema 引用的 data_key 在 data dict 中缺失
    """
    if not schema_yaml.exists():
        raise FileNotFoundError(f"Schema YAML missing: {schema_yaml}")

    schema = yaml.safe_load(schema_yaml.read_text(encoding="utf-8"))
    sheets_def = schema.get("sheets", [])
    if not sheets_def:
        raise ValueError("Schema must define at least one sheet under 'sheets'")

    wb = Workbook()
    wb.remove(wb.active)  # 删默认 sheet

    sheet_names = []
    for idx, sheet_def in enumerate(sheets_def):
        if "name" not in sheet_def:
            raise ValueError(f"Sheet #{idx} missing required field 'name'")
        if "headers" not in sheet_def:
            raise ValueError(f"Sheet #{idx} ('{sheet_def.get('name', '?')}') missing required field 'headers'")
        if "data_key" not in sheet_def:
            raise ValueError(f"Sheet #{idx} ('{sheet_def.get('name', '?')}') missing required field 'data_key'")

        name = sheet_def["name"]
        headers = sheet_def["headers"]
        widths = sheet_def.get("column_widths", [16] * len(headers))
        data_key = sheet_def["data_key"]

        if data_key not in data:
            raise KeyError(f"data['{data_key}'] missing for sheet '{name}'")

        ws = wb.create_sheet(title=name)
        ws.append(headers)
        for row in data[data_key]:
            ws.append(row)

        _style_header(ws)
        _style_body(ws)
        _autofit(ws, widths)
        sheet_names.append(name)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    validation = _validate_xlsx(out_path, len(sheet_names)) if auto_validate else None
    ok = (validation is None) or (validation["size_ok"] and validation["sheet_count_ok"])

    return {
        "ok": ok,
        "out_path": str(out_path),
        "size_mb": round(out_path.stat().st_size / (1024 * 1024), 4),
        "sheets": sheet_names,
        "validation": validation,
    }
